from cgitb import text
from email import message
from msilib.schema import ComboBox
from os import name
from re import L
from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from turtle import width
from openpyxl.workbook import Workbook
from openpyxl import load_workbook


root = tk.Tk()
root.geometry('260x130')
root.title('LogIN')
root.configure(bg='grey')

excel_con = Workbook()
excel_con = load_workbook('xlsx.xlsx')
excel_activate = excel_con.active

column1 = excel_activate['A']
column2 = excel_activate['B']
column3 = excel_activate['C']
column4 = excel_activate['D']
column5 = excel_activate['E']
column6 = excel_activate['F']
column7 = excel_activate['G']
column8 = excel_activate['H']

for every_cell in column1:
    print(every_cell.value)

for every_cell in column2:
    print(every_cell.value)

for every_cell in column3:
    print(every_cell.value)

for every_cell in column4:
    print(every_cell.value)

for every_cell in column5:
    print(every_cell.value)

for every_cell in column6:
    print(every_cell.value)

for every_cell in column7:
    print(every_cell.value)

for every_cell in column8:
    print(every_cell.value)



def loginfunction():
    if username.get() == "admin" and password.get() == "123":
        messagebox.showinfo("Login ",">>> LOGIN SUCCESSFULLY <<<")
        root.withdraw()
        newinterface()
    else:
        messagebox.showerror("Please Try Again","LOGIN DENIED")

def cancelfunction():
    root.destroy()


user_label = Label(root,text="USERNAME:",width=10,bg='dimgray',font=('courier'),fg='white')
pass_label = Label(root,text="PASSWORD:",width=10,bg='dimgray',font=('courier'),fg='white')
username = Entry(root,width=20)
password = Entry(root,width=20,show="•")
login = Button(root,text="LOGIN",width=10,command=lambda:loginfunction(),bg='dimgray',font=('courier'),fg='white')
cancel = Button(root,text="Cancel",width=10,command=lambda:cancelfunction(),bg='dimgray',font=('courier'),fg='white')


user_label.grid(row=1,column=0,padx=5,pady=5)
pass_label.grid(row=2,column=0,padx=5,pady=5)
username.grid(row=1,column=1,padx=5,pady=5)
password.grid(row=2,column=1,padx=5,pady=5)
login.grid(row=3,column=0,padx=5,pady=5)
cancel.grid(row=3,column=1,padx=5,pady=5,sticky=S)

def newinterface():
    groot = tk.Toplevel()
    groot.geometry('1200x760')
    groot.title('NewInterface')
    groot.configure(bg='dimgray')
    
    messagebox.showinfo("ERS(Employee Record System) ","Welcome To ERS")
    
    def search():
        Search = tk.Toplevel()
        Search.geometry('800x350')
        Search.title('SEARCH')
        Search.configure(bg = 'grey')

        def search_name():
            Found = True
            s_name = name_ntr.get()
            for each_cell in range(2, excel_activate.max_row + 1):
                if excel_activate['B'+str(each_cell)].value == s_name:
                    messagebox.showinfo("DATA",f'Data Found @ Record Number{each_cell}')
                    Found = False
                    break
            if Found == True:
                messagebox.showerror("DATA",'Data Not Found')
            
                
        def exit_function1():
            Search.destroy()

        question_label = Label(Search,text='Input An Name',width=30,bg='gray',fg='white',font=('courier',23,'bold'),borderwidth=4, relief="groove")
        question_label.grid(row=0,column=1,padx=10,pady=10)

        name_ntr = Entry(Search,width=30,bg='gray',fg='white',font=('courier',23,'bold'),borderwidth=4, relief="groove")
        name_ntr.grid(row=1,column=1,padx=10,pady=10)

        search_btn = Button(Search,text='Search',width=30,bg='gray',fg='white',font=('courier',23,'bold'),borderwidth=4, relief="groove",command=lambda:search_name())
        search_btn.grid(row=2,column=1,padx=10,pady=10)
            
        exit_btn = Button(Search,text='Exit',width=30,bg='gray',fg='white',font=('courier',23,'bold'),borderwidth=4, relief="groove",command=lambda:exit_function1())
        exit_btn.grid(row=3,column=1,padx=10,pady=10)

    def save_record():
        ref = ref_entry.get()
        name = name_entry.get()
        email = em_entry.get()
        gender = gender_var.get()
        designation = designation_var.get()
        contact = contact_entry.get()
        salary = salary_entry.get()
        address = add_text.get(1.0,END)
        Found = False

        for every_cell in range(2, excel_activate.max_row + 1):
            if ref == excel_activate['A'+str(every_cell)].value:
                Found = True
                break
        if Found == True:
            messagebox.showerror('DATA','Data Already Exist')
        else:
            lastrow = str((excel_activate.max_row)+1)
            excel_activate['A'+lastrow] = ref
            excel_activate['B'+lastrow] = name
            excel_activate['C'+lastrow] = email
            excel_activate['D'+lastrow] = gender
            excel_activate['E'+lastrow] = designation
            excel_activate['F'+lastrow] = contact
            excel_activate['G'+lastrow] = salary
            excel_activate['H'+lastrow] = address
            
            excel_con.save('xlsx.xlsx')
            messagebox.showinfo("Save Record","Saved Successfully")


    
    def reset_record():
        ref_entry.delete(0,END)
        name_entry.delete(0,END)
        em_entry.delete(0,END)
        gender_var.set("")
        designation_var.set("")
        contact_entry.delete(0,END)
        salary_entry.delete(0,END)
        add_text.delete("1.0","end-1c")
        print_text.delete("1.0","end")
        

    def print_record():
        ref = ref_entry.get()
        name = name_entry.get()
        email = em_entry.get()
        gender = gender_var.get()
        designation = designation_var.get()
        contact = contact_entry.get()
        salary = salary_entry.get()
        address = add_text.get("1.0","end")
        
        print_text.insert(END,f'Employee Ref :\t{ref}\nFull Name :\t{name}\nEmails :\t{email}\nGender :\t{gender}\nDesignation :\t{designation}\nContact No. :\t{contact}\nSalary :\t{salary}\nAdress :\t{address}')
        
        messagebox.showinfo("Printed Record","Print Successfully")

    def exit_function():
        groot.destroy()

    
     
    title_label = Label(groot,text="  Employee Records System  ",width=40,height=1,font=('courier',36,'bold'),bg='gray',fg='white',borderwidth=20, relief="groove")
    title_label.place(x=0,y=0)
        
    frame1 = Frame(groot,width=600,height=640,bg='dimgray',borderwidth=8, relief="groove")
    frame1.place(x=0,y=95)

    fillup_label = Label(frame1,text='Fill Up Form',width=30,bg='gray',fg='white',font=('courier',24,'bold'),borderwidth=8, relief="groove")
    fillup_label.grid(row=0,column=0,columnspan=3)

    ref_label = Label(frame1,text="Employee Ref : ",width=20,bg="dimgray",fg='white',font=('courier',15,'bold'))
    ref_entry = Entry(frame1,width=28,font=('courier'))

    ref_label.grid(row=1,column=0,pady=10)
    ref_entry.grid(row=1,column=1,padx=10,pady=10)

    name_label = Label(frame1,text="Full Name : ",width=20,bg="dimgray",fg='white',font=('courier',15,'bold'))
    name_entry = Entry(frame1,width=28,font=('courier'))

    name_label.grid(row=2,column=0,pady=10)
    name_entry.grid(row=2,column=1,pady=10)

    em_label = Label(frame1,text="Emails : ",width=20,bg="dimgray",fg='white',font=('courier',15,'bold'))
    em_entry = Entry(frame1,width=28,font=('courier'))


    em_label.grid(row=3,column=0,pady=10)
    em_entry.grid(row=3,column=1,pady=10)

    gender_label = Label(frame1,text="Gender :",width=10,bg="dimgray",fg='white',font=('courier',15,'bold'))
    gender_var = StringVar()
    gender_list = ["Male","Female","Other","Rather Not Say"]
    gender_combo = ttk.Combobox(frame1,textvariable=gender_var,value=gender_list,width=27,font=('courier'))

    gender_label.grid(row=4,column=0,pady=11)
    gender_combo.grid(row=4,column=1,pady=11)

    designation_label = Label(frame1,text="Designation :",width=13,bg="dimgray",fg='white',font=('courier',15,'bold'))
    designation_var = StringVar()
    designation_list = ["IT","Security","Office","Cleaning Crew"]
    designation_combo = ttk.Combobox(frame1,textvariable=designation_var,value=designation_list,width=27,font=('courier'))

    designation_label.grid(row=5,column=0,pady=11)
    designation_combo.grid(row=5,column=1,pady=12)

    contact_label = Label(frame1,text="Contact No. : ",width=20,bg="dimgray",fg='white',font=('courier',15,'bold'))
    contact_entry = Entry(frame1,width=28,font=('courier'))

    contact_label.grid(row=6,column=0,pady=12)
    contact_entry.grid(row=6,column=1,pady=12)

    salary_label = Label(frame1,text="Salary : ",width=20,bg="dimgray",fg='white',font=('courier',15,'bold'))
    salary_entry = Entry(frame1,width=28,font=('courier'))

    salary_label.grid(row=7,column=0,pady=12)
    salary_entry.grid(row=7,column=1,pady=12)

    add_label = Label(frame1,text="Address : ",width=20,bg="dimgray",fg='white',font=('courier',15,'bold'))
    add_text = Text(frame1,width=28,height=4,font=('courier'))

    add_label.grid(row=8,column=0,pady=12)
    add_text.grid(row=8,column=1,pady=12)

    frame2 = Frame(groot,width=580,height=640,bg='dimgray',borderwidth=8, relief="groove")
    frame2.place(x=605,y=95)

    detail_label = Label(frame2,text='Employees Details',width=30,bg='gray',fg='white',font=('courier',23,'bold'),borderwidth=4, relief="groove")
    detail_label.grid(row=0,column=0,columnspan=5)

    print_text = Text(frame2,width=35,height=8,font=('courier',20,'bold'),bg='dimgray',fg='white')
    print_text.grid(row=1,column=0,columnspan=5,pady=10)

    global tv
    tv = ttk.Treeview(frame2,height=8)
    treescrolly = Scrollbar(frame2, orient="vertical", command= tv.yview)
    treescrollx = Scrollbar(frame2, orient="horizontal", command= tv.xview)
    tv.configure(xscrollcommand = treescrollx.set, yscrollcommand=treescrolly.set) 
    treescrolly.grid(row=2,column=4)
    treescrollx.grid(row=3,column=0,columnspan=4)
    tv['columns'] = ("Ref", "Full Name", "Email","Gender","Designation")
    tv.column("Ref", anchor=W, width=60)
    tv.column("Full Name",  anchor=CENTER, width=120)
    tv.column("Email", anchor=W, width=120)
    tv.column("Gender", anchor=W, width=60)
    tv.column("Designation", anchor=W, width=60)

    tv.heading("Ref", text="Ref", anchor=W)
    tv.heading("Full Name", text="Full Name", anchor=CENTER)
    tv.heading("Email", text="Email", anchor=W)
    tv.heading("Gender", text="Gender", anchor=W)
    tv.heading("Designation", text="Designation", anchor=W)

    for each_cell in range(2, (excel_activate.max_row)+1):
                                                    #NAME                                           #PHONE                                               #ADDRESS       
        tv.insert(parent='', index="end", values=(excel_activate['A'+str(each_cell)].value,excel_activate['B'+str(each_cell)].value, excel_activate['C'+str(each_cell)].value, excel_activate['D'+str(each_cell)].value, excel_activate['E'+str(each_cell)].value))
    tv.grid(row=2,column=0,columnspan=4)

    frame3 = Frame(groot,width=1300,height=60,bg='gray',borderwidth=8, relief="groove")
    frame3.place(x=0,y=620)

    search_button = Button(frame3,text='Search',width=15,font=('courier',15,'bold'),height=2,command=lambda:search())
    search_button.grid(row=0,column=0,padx=24,pady=30)

    save_button = Button(frame3,text='Save',width=15,font=('courier',15,'bold'),height=2,command=lambda:save_record())
    save_button.grid(row=0,column=1,padx=22,pady=30)

    print_button = Button(frame3,text='Print',width=15,font=('courier',15,'bold'),height=2,command=lambda:print_record())
    print_button.grid(row=0,column=2,padx=24,pady=30)

    reset_button = Button(frame3,text='Reset',width=15,font=('courier',15,'bold'),height=2,command=lambda:reset_record())
    reset_button.grid(row=0,column=3,padx=24,pady=30)

    exit_button = Button(frame3,text='Exit',width=15,font=('courier',15,'bold'),command=lambda:exit_function(),height=2)
    exit_button.grid(row=0,column=4,padx=22,pady=30)
    groot.mainloop()


root.mainloop()

